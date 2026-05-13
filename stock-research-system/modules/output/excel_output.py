#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel输出模块 - 生成带公式的专业Excel财务模型
参考Claude Financial Services的xlsx-author技能实现
"""

import logging
from typing import Dict, Any, List, Optional
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference

logger = logging.getLogger(__name__)


class ExcelOutput:
    """Excel输出基类"""
    
    def __init__(self, config: Dict[str, Any]):
        """
        初始化
        
        Args:
            config: 系统配置
        """
        self.config = config
        self.output_path = Path(config['data_sources'].get('report_output_path', 'output/reports'))
        self.output_path.mkdir(parents=True, exist_ok=True)
        
        # 样式定义
        self._init_styles()
        
    def _init_styles(self):
        """初始化Excel样式"""
        # 标题样式
        self.title_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        self.title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        # 表头样式
        self.header_font = Font(name='Arial', size=11, bold=True)
        self.header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        
        # 数字样式
        self.number_font = Font(name='Arial', size=10)
        
        # 边框
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
    def _apply_title_style(self, cell):
        """应用标题样式"""
        cell.font = self.title_font
        cell.fill = self.title_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
    def _apply_header_style(self, cell):
        """应用表头样式"""
        cell.font = self.header_font
        cell.fill = self.header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = self.thin_border
        
    def _apply_number_style(self, cell):
        """应用数字样式"""
        cell.font = self.number_font
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = self.thin_border
        
    def _set_column_width(self, ws, col, width):
        """设置列宽"""
        ws.column_dimensions[get_column_letter(col)].width = width
        
    def _write_title(self, ws, row, col, title, merge_to_col=None):
        """写入标题"""
        cell = ws.cell(row=row, column=col, value=title)
        self._apply_title_style(cell)
        
        if merge_to_col:
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to_col)
            
    def _write_header(self, ws, row, col, value):
        """写入表头"""
        cell = ws.cell(row=row, column=col, value=value)
        self._apply_header_style(cell)
        
    def _write_number(self, ws, row, col, value, number_format='#,##0'):
        """写入数字"""
        cell = ws.cell(row=row, column=col, value=value)
        self._apply_number_style(cell)
        cell.number_format = number_format
        
    def _write_formula(self, ws, row, col, formula, number_format='#,##0'):
        """写入公式"""
        cell = ws.cell(row=row, column=col, value=formula)
        self._apply_number_style(cell)
        cell.number_format = number_format
        
    def save(self, wb: Workbook, filename: str) -> str:
        """
        保存Excel文件
        
        Args:
            wb: Workbook对象
            filename: 文件名
            
        Returns:
            文件路径
        """
        filepath = self.output_path / filename
        wb.save(filepath)
        logger.info(f"Excel文件已保存: {filepath}")
        return str(filepath)


class DCFExcelOutput(ExcelOutput):
    """DCF模型Excel输出"""
    
    def generate(self, dcf_result: Dict[str, Any], stock_code: str) -> str:
        """
        生成DCF模型Excel文件
        
        Args:
            dcf_result: DCF分析结果
            stock_code: 股票代码
            
        Returns:
            文件路径
        """
        logger.info(f"生成DCF模型Excel: {stock_code}")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "DCF模型"
        
        # 标题
        self._write_title(ws, 1, 1, f"DCF估值模型 - {stock_code}", merge_to_col=6)
        
        # 假设部分
        row = 3
        self._write_title(ws, row, 1, "关键假设", merge_to_col=2)
        
        row += 1
        assumptions = dcf_result.get('assumptions', {})
        
        headers = ['参数', '数值', '单位']
        for col, header in enumerate(headers, 1):
            self._write_header(ws, row, col, header)
            
        row += 1
        data_rows = [
            ('预测期', assumptions.get('forecast_years', 5), '年'),
            ('永续增长率', assumptions.get('growth_rate', 0.05), '%'),
            ('WACC', assumptions.get('wacc', 0.10), '%'),
            ('基准自由现金流', dcf_result.get('base_fcf', 0), '元'),
        ]
        
        for item in data_rows:
            self._write_header(ws, row, 1, item[0])
            if isinstance(item[1], float) and item[2] == '%':
                self._write_number(ws, row, 2, item[1], '0.00%')
            else:
                self._write_number(ws, row, 2, item[1])
            self._write_header(ws, row, 3, item[2])
            row += 1
            
        # 现金流预测
        row += 2
        self._write_title(ws, row, 1, "自由现金流预测", merge_to_col=6)
        
        row += 1
        forecast_headers = ['年份', '增长率', '自由现金流', '折现因子', '现值']
        for col, header in enumerate(forecast_headers, 1):
            self._write_header(ws, row, col, header)
            
        row += 1
        forecasted_fcf = dcf_result.get('forecasted_fcf', [])
        wacc = assumptions.get('wacc', 0.10)
        growth_rate = assumptions.get('growth_rate', 0.05)
        
        start_row = row
        for i, fcf in enumerate(forecasted_fcf, 1):
            self._write_number(ws, row, 1, i)
            
            # 增长率（逐年递减）
            year_growth = growth_rate * (1 - (i-1) * 0.1)
            self._write_number(ws, row, 2, year_growth, '0.00%')
            
            # 自由现金流
            self._write_number(ws, row, 3, fcf, '#,##0')
            
            # 折现因子
            discount_factor = f"=1/(1+$D$6)^{i}"  # 引用WACC单元格
            self._write_formula(ws, row, 4, discount_factor, '0.0000')
            
            # 现值
            self._write_formula(ws, row, 5, f"=C{row}*D{row}", '#,##0')
            
            row += 1
            
        # 终值计算
        row += 1
        self._write_header(ws, row, 1, "终值计算")
        
        last_fcf_row = row - 2
        terminal_value_formula = f"=C{last_fcf_row}*(1+$D$7)/($D$6-$D$7)"  # FCF*(1+g)/(WACC-g)
        self._write_formula(ws, row, 3, terminal_value_formula, '#,##0')
        
        # 终值现值
        row += 1
        self._write_header(ws, row, 1, "终值现值")
        forecast_years = assumptions.get('forecast_years', 5)
        terminal_pv_formula = f"=C{row-1}/(1+$D$6)^{forecast_years}"
        self._write_formula(ws, row, 3, terminal_pv_formula, '#,##0')
        
        # 企业价值
        row += 2
        self._write_title(ws, row, 1, "估值结果", merge_to_col=2)
        
        row += 1
        self._write_header(ws, row, 1, "预测期现金流现值")
        pv_range = f"=SUM(E{start_row}:E{start_row + len(forecasted_fcf) - 1})"
        self._write_formula(ws, row, 2, pv_range, '#,##0')
        
        row += 1
        self._write_header(ws, row, 1, "终值现值")
        self._write_formula(ws, row, 2, f"=C{row-3}", '#,##0')
        
        row += 1
        self._write_header(ws, row, 1, "企业价值")
        self._write_formula(ws, row, 2, f"=B{row-2}+B{row-1}", '#,##0')
        
        row += 1
        self._write_header(ws, row, 1, "净债务")
        net_debt = dcf_result.get('net_debt', 0)
        self._write_number(ws, row, 2, net_debt, '#,##0')
        
        row += 1
        self._write_header(ws, row, 1, "股权价值")
        self._write_formula(ws, row, 2, f"=B{row-2}-B{row-1}", '#,##0')
        
        row += 1
        self._write_header(ws, row, 1, "流通股数（亿股）")
        shares = dcf_result.get('shares_outstanding', 10)
        self._write_number(ws, row, 2, shares, '0.00')
        
        row += 1
        self._write_header(ws, row, 1, "每股价值（元）")
        self._write_formula(ws, row, 2, f"=B{row-2}/(B{row-1}*100000000)", '0.00')
        
        # 敏感性分析
        sensitivity = dcf_result.get('sensitivity_analysis', {})
        if sensitivity:
            row += 2
            self._write_title(ws, row, 1, "敏感性分析", merge_to_col=6)
            
            row += 1
            self._write_header(ws, row, 1, "增长率\\WACC")
            
            wacc_range = sensitivity.get('wacc_range', [])
            for col, wacc_val in enumerate(wacc_range, 2):
                self._write_header(ws, row, col, f"{wacc_val:.2%}")
                
            growth_range = sensitivity.get('growth_range', [])
            matrix = sensitivity.get('matrix', [])
            
            for i, growth_val in enumerate(growth_range):
                row += 1
                self._write_header(ws, row, 1, f"{growth_val:.2%}")
                
                if i < len(matrix):
                    for col, value in enumerate(matrix[i], 2):
                        self._write_number(ws, row, col, value, '0.00')
                        
        # 设置列宽
        self._set_column_width(ws, 1, 20)
        self._set_column_width(ws, 2, 15)
        self._set_column_width(ws, 3, 15)
        self._set_column_width(ws, 4, 12)
        self._set_column_width(ws, 5, 15)
        
        # 保存文件
        filename = f"{stock_code}_DCF模型_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return self.save(wb, filename)


class CompsExcelOutput(ExcelOutput):
    """可比公司分析Excel输出"""
    
    def generate(self, comps_result: Dict[str, Any], stock_code: str) -> str:
        """
        生成可比公司分析Excel文件
        
        Args:
            comps_result: 可比公司分析结果
            stock_code: 股票代码
            
        Returns:
            文件路径
        """
        logger.info(f"生成可比公司分析Excel: {stock_code}")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "可比公司分析"
        
        # 标题
        self._write_title(ws, 1, 1, f"可比公司分析 - {stock_code}", merge_to_col=6)
        
        # 表头
        row = 3
        headers = ['公司', 'PE', 'PB', 'EV/EBITDA', 'PEG', '备注']
        for col, header in enumerate(headers, 1):
            self._write_header(ws, row, col, header)
            
        # 目标公司
        row += 1
        target = comps_result.get('target_company', {})
        multiples = comps_result.get('valuation_multiples', {})
        
        self._write_header(ws, row, 1, f"**{target.get('stock_name', stock_code)}**")
        
        target_multiples = multiples.get('target', {})
        self._write_number(ws, row, 2, target_multiples.get('PE', 'N/A'), '0.0')
        self._write_number(ws, row, 3, target_multiples.get('PB', 'N/A'), '0.0')
        self._write_number(ws, row, 4, target_multiples.get('EV_EBITDA', 'N/A'), '0.0')
        self._write_number(ws, row, 5, target_multiples.get('PEG', 'N/A'), '0.00')
        self._write_header(ws, row, 6, '目标公司')
        
        # 可比公司
        peers = comps_result.get('peer_companies', [])
        for peer in peers:
            row += 1
            self._write_header(ws, row, 1, peer.get('stock_name', peer.get('stock_code', 'N/A')))
            self._write_number(ws, row, 2, peer.get('pe_ratio', 'N/A'), '0.0')
            self._write_number(ws, row, 3, peer.get('pb_ratio', 'N/A'), '0.0')
            self._write_number(ws, row, 4, peer.get('ev_ebitda', 'N/A'), '0.0')
            self._write_number(ws, row, 5, peer.get('peg_ratio', 'N/A'), '0.00')
            self._write_header(ws, row, 6, '可比公司')
            
        # 行业平均
        row += 1
        peer_avg = multiples.get('peer_average', {})
        self._write_header(ws, row, 1, '**行业平均**')
        self._write_number(ws, row, 2, peer_avg.get('PE', 'N/A'), '0.0')
        self._write_number(ws, row, 3, peer_avg.get('PB', 'N/A'), '0.0')
        self._write_number(ws, row, 4, peer_avg.get('EV_EBITDA', 'N/A'), '0.0')
        self._write_number(ws, row, 5, peer_avg.get('PEG', 'N/A'), '0.00')
        self._write_header(ws, row, 6, 'AVERAGE公式')
        
        # 添加AVERAGE公式（如果有可比公司）
        if peers:
            start_row = 5
            end_row = 4 + len(peers)
            
            ws.cell(row=row, column=2).value = f"=AVERAGE(B{start_row}:B{end_row})"
            ws.cell(row=row, column=3).value = f"=AVERAGE(C{start_row}:C{end_row})"
            ws.cell(row=row, column=4).value = f"=AVERAGE(D{start_row}:D{end_row})"
            ws.cell(row=row, column=5).value = f"=AVERAGE(E{start_row}:E{end_row})"
            
        # 估值差距
        row += 2
        self._write_title(ws, row, 1, "估值差距分析", merge_to_col=3)
        
        row += 1
        self._write_header(ws, row, 1, "指标")
        self._write_header(ws, row, 2, "差距")
        self._write_header(ws, row, 3, "说明")
        
        valuation_gap = multiples.get('valuation_gap', {})
        for metric, gap in valuation_gap.items():
            row += 1
            self._write_header(ws, row, 1, metric)
            self._write_number(ws, row, 2, gap, '0.0%')
            
            if gap < -0.2:
                assessment = "显著低估"
            elif gap < -0.1:
                assessment = "低估"
            elif gap < 0.1:
                assessment = "合理"
            elif gap < 0.2:
                assessment = "略高"
            else:
                assessment = "显著高估"
            self._write_header(ws, row, 3, assessment)
            
        # 隐含估值
        row += 2
        self._write_title(ws, row, 1, "隐含估值", merge_to_col=3)
        
        row += 1
        self._write_header(ws, row, 1, "方法")
        self._write_header(ws, row, 2, "隐含价值（元/股）")
        
        implied = comps_result.get('implied_valuation', {})
        
        row += 1
        self._write_header(ws, row, 1, "基于PE")
        self._write_number(ws, row, 2, implied.get('by_pe', 0), '0.00')
        
        row += 1
        self._write_header(ws, row, 1, "基于PB")
        self._write_number(ws, row, 2, implied.get('by_pb', 0), '0.00')
        
        row += 1
        self._write_header(ws, row, 1, "基于EV/EBITDA")
        self._write_number(ws, row, 2, implied.get('by_ev_ebitda', 0), '0.00')
        
        row += 1
        self._write_header(ws, row, 1, "平均隐含价值")
        self._write_formula(ws, row, 2, f"=AVERAGE(B{row-3}:B{row-1})", '0.00')
        
        # 评估
        row += 2
        self._write_title(ws, row, 1, "综合评估", merge_to_col=3)
        
        row += 1
        assessment = comps_result.get('assessment', '')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.cell(row=row, column=1, value=assessment)
        
        # 设置列宽
        self._set_column_width(ws, 1, 20)
        self._set_column_width(ws, 2, 15)
        self._set_column_width(ws, 3, 15)
        self._set_column_width(ws, 4, 15)
        self._set_column_width(ws, 5, 12)
        self._set_column_width(ws, 6, 15)
        
        # 保存文件
        filename = f"{stock_code}_可比公司分析_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return self.save(wb, filename)


class FinancialStatementExcelOutput(ExcelOutput):
    """三大财务报表Excel输出"""
    
    def generate(self, financial_result: Dict[str, Any], stock_code: str) -> str:
        """
        生成三大财务报表Excel文件
        
        Args:
            financial_result: 财务分析结果
            stock_code: 股票代码
            
        Returns:
            文件路径
        """
        logger.info(f"生成财务报表Excel: {stock_code}")
        
        wb = Workbook()
        
        # 资产负债表
        self._create_balance_sheet(wb, financial_result, stock_code)
        
        # 利润表
        self._create_income_statement(wb, financial_result, stock_code)
        
        # 现金流量表
        self._create_cash_flow(wb, financial_result, stock_code)
        
        # 财务指标
        self._create_financial_indicators(wb, financial_result, stock_code)
        
        # 保存文件
        filename = f"{stock_code}_财务报表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return self.save(wb, filename)
        
    def _create_balance_sheet(self, wb: Workbook, financial_result: Dict, stock_code: str):
        """创建资产负债表"""
        ws = wb.active
        ws.title = "资产负债表"
        
        # 标题
        self._write_title(ws, 1, 1, f"资产负债表 - {stock_code}", merge_to_col=5)
        
        # 表头
        row = 3
        headers = ['项目', '2024', '2023', '2022', '2021']
        for col, header in enumerate(headers, 1):
            self._write_header(ws, row, col, header)
            
        # 数据
        bs_analysis = financial_result.get('balance_sheet_analysis', {})
        total_assets_trend = bs_analysis.get('total_assets_trend', [])
        
        # 总资产
        row += 1
        self._write_header(ws, row, 1, "资产总计")
        for i, item in enumerate(total_assets_trend[:4], 2):
            self._write_number(ws, row, i, item.get('资产总计', 0), '#,##0')
            
        # 关键比率
        row += 2
        self._write_title(ws, row, 1, "关键比率", merge_to_col=2)
        
        row += 1
        key_ratios = bs_analysis.get('key_ratios', {})
        
        ratio_items = [
            ('资产负债率', key_ratios.get('debt_to_asset', 0), '0.00%'),
            ('流动比率', key_ratios.get('current_ratio', 0), '0.00'),
            ('速动比率', key_ratios.get('quick_ratio', 0), '0.00'),
        ]
        
        for item in ratio_items:
            self._write_header(ws, row, 1, item[0])
            self._write_number(ws, row, 2, item[1], item[2])
            row += 1
            
        # 设置列宽
        self._set_column_width(ws, 1, 25)
        for col in range(2, 6):
            self._set_column_width(ws, col, 15)
            
    def _create_income_statement(self, wb: Workbook, financial_result: Dict, stock_code: str):
        """创建利润表"""
        ws = wb.create_sheet("利润表")
        
        # 标题
        self._write_title(ws, 1, 1, f"利润表 - {stock_code}", merge_to_col=5)
        
        # 表头
        row = 3
        headers = ['项目', '2024', '2023', '2022', '2021']
        for col, header in enumerate(headers, 1):
            self._write_header(ws, row, col, header)
            
        # 数据
        is_analysis = financial_result.get('income_statement_analysis', {})
        revenue_trend = is_analysis.get('revenue_trend', [])
        profit_trend = is_analysis.get('profit_trend', [])
        
        # 营业收入
        row += 1
        self._write_header(ws, row, 1, "营业总收入")
        for i, item in enumerate(revenue_trend[:4], 2):
            self._write_number(ws, row, i, item.get('营业总收入', 0), '#,##0')
            
        # 净利润
        row += 1
        self._write_header(ws, row, 1, "净利润")
        for i, item in enumerate(profit_trend[:4], 2):
            self._write_number(ws, row, i, item.get('净利润', 0), '#,##0')
            
        # 利润率
        row += 2
        self._write_title(ws, row, 1, "利润率分析", merge_to_col=2)
        
        row += 1
        margins = is_analysis.get('margin_analysis', {})
        
        margin_items = [
            ('毛利率', margins.get('gross_margin', 0), '0.00%'),
            ('营业利润率', margins.get('operating_margin', 0), '0.00%'),
            ('净利率', margins.get('net_margin', 0), '0.00%'),
        ]
        
        for item in margin_items:
            self._write_header(ws, row, 1, item[0])
            self._write_number(ws, row, 2, item[1], item[2])
            row += 1
            
        # 设置列宽
        self._set_column_width(ws, 1, 25)
        for col in range(2, 6):
            self._set_column_width(ws, col, 15)
            
    def _create_cash_flow(self, wb: Workbook, financial_result: Dict, stock_code: str):
        """创建现金流量表"""
        ws = wb.create_sheet("现金流量表")
        
        # 标题
        self._write_title(ws, 1, 1, f"现金流量表 - {stock_code}", merge_to_col=5)
        
        # 表头
        row = 3
        headers = ['项目', '2024', '2023', '2022', '2021']
        for col, header in enumerate(headers, 1):
            self._write_header(ws, row, col, header)
            
        # 数据
        cf_analysis = financial_result.get('cash_flow_analysis', {})
        ocf_trend = cf_analysis.get('operating_cf_trend', [])
        
        # 经营现金流
        row += 1
        self._write_header(ws, row, 1, "经营活动现金流净额")
        for i, item in enumerate(ocf_trend[:4], 2):
            self._write_number(ws, row, i, item.get('经营活动产生的现金流量净额', 0), '#,##0')
            
        # 现金流质量
        row += 2
        self._write_title(ws, row, 1, "现金流质量", merge_to_col=2)
        
        row += 1
        quality = cf_analysis.get('quality_analysis', {})
        
        quality_items = [
            ('经营现金流/净利润', quality.get('ocf_to_net_profit', 0), '0.00'),
            ('自由现金流', quality.get('free_cash_flow', 0), '#,##0'),
        ]
        
        for item in quality_items:
            self._write_header(ws, row, 1, item[0])
            self._write_number(ws, row, 2, item[1], item[2])
            row += 1
            
        # 设置列宽
        self._set_column_width(ws, 1, 25)
        for col in range(2, 6):
            self._set_column_width(ws, col, 15)
            
    def _create_financial_indicators(self, wb: Workbook, financial_result: Dict, stock_code: str):
        """创建财务指标汇总"""
        ws = wb.create_sheet("财务指标汇总")
        
        # 标题
        self._write_title(ws, 1, 1, f"财务指标汇总 - {stock_code}", merge_to_col=3)
        
        # 表头
        row = 3
        headers = ['指标类别', '指标名称', '数值', '评估']
        for col, header in enumerate(headers, 1):
            self._write_header(ws, row, col, header)
            
        # 汇总数据
        summary = financial_result.get('summary', {})
        key_metrics = summary.get('key_metrics', {})
        strengths = summary.get('strengths', [])
        weaknesses = summary.get('weaknesses', [])
        
        # 盈利能力
        row += 1
        self._write_header(ws, row, 1, "盈利能力")
        
        row += 1
        self._write_header(ws, row, 2, "毛利率")
        self._write_number(ws, row, 3, key_metrics.get('gross_margin', 'N/A'), '0.00%')
        self._write_header(ws, row, 4, "良好" if key_metrics.get('gross_margin', 0) > 0.3 else "一般")
        
        row += 1
        self._write_header(ws, row, 2, "净利率")
        self._write_number(ws, row, 3, key_metrics.get('net_margin', 'N/A'), '0.00%')
        self._write_header(ws, row, 4, "良好" if key_metrics.get('net_margin', 0) > 0.1 else "一般")
        
        # 偿债能力
        row += 1
        self._write_header(ws, row, 1, "偿债能力")
        
        row += 1
        self._write_header(ws, row, 2, "资产负债率")
        self._write_number(ws, row, 3, key_metrics.get('debt_to_asset', 'N/A'), '0.00%')
        debt_ratio = key_metrics.get('debt_to_asset', 0)
        self._write_header(ws, row, 4, "稳健" if debt_ratio < 0.5 else "需关注")
        
        row += 1
        self._write_header(ws, row, 2, "流动比率")
        self._write_number(ws, row, 3, key_metrics.get('current_ratio', 'N/A'), '0.00')
        current_ratio = key_metrics.get('current_ratio', 0)
        self._write_header(ws, row, 4, "良好" if current_ratio > 1.5 else "一般")
        
        # 现金流质量
        row += 1
        self._write_header(ws, row, 1, "现金流质量")
        
        row += 1
        self._write_header(ws, row, 2, "经营现金流/净利润")
        self._write_number(ws, row, 3, key_metrics.get('ocf_to_net_profit', 'N/A'), '0.00')
        ocf_ratio = key_metrics.get('ocf_to_net_profit', 0)
        self._write_header(ws, row, 4, "优秀" if ocf_ratio > 1.2 else "良好" if ocf_ratio > 0.8 else "需关注")
        
        # 设置列宽
        self._set_column_width(ws, 1, 15)
        self._set_column_width(ws, 2, 20)
        self._set_column_width(ws, 3, 15)
        self._set_column_width(ws, 4, 15)
