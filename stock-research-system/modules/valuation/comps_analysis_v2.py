#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
可比公司分析模块 - 基于Claude Financial Services框架
包括：运营指标、估值倍数、统计分析、Excel输出
"""

import logging
from typing import Dict, Any, List, Optional
import pandas as pd
import numpy as np
from datetime import datetime
from pathlib import Path

logger = logging.getLogger(__name__)


class CompsAnalyzerV2:
    """可比公司分析器 - Claude框架"""
    
    def __init__(self, config: Dict[str, Any]):
        """
        初始化可比公司分析器
        
        Args:
            config: 系统配置
        """
        self.config = config
        
        # 行业特定指标配置
        self.industry_metrics = {
            'SaaS': {
                'operating_metrics': ['Revenue', 'Revenue Growth', 'Gross Margin', 'EBITDA Margin', 'Rule of 40'],
                'valuation_metrics': ['Market Cap', 'EV', 'EV/Revenue', 'EV/EBITDA', 'P/E'],
            },
            'Manufacturing': {
                'operating_metrics': ['Revenue', 'Revenue Growth', 'Gross Margin', 'EBITDA Margin', 'Asset Turnover'],
                'valuation_metrics': ['Market Cap', 'EV', 'EV/EBITDA', 'P/E', 'P/B'],
            },
            'Financial Services': {
                'operating_metrics': ['Revenue', 'Revenue Growth', 'ROE', 'ROA', 'Net Interest Margin'],
                'valuation_metrics': ['Market Cap', 'P/E', 'P/B', 'Dividend Yield'],
            },
            'Default': {
                'operating_metrics': ['Revenue', 'Revenue Growth', 'Gross Margin', 'EBITDA Margin', 'Net Margin'],
                'valuation_metrics': ['Market Cap', 'EV', 'EV/Revenue', 'EV/EBITDA', 'P/E'],
            },
        }
    
    def analyze(
        self,
        target_company: Dict[str, Any],
        peer_companies: List[Dict[str, Any]],
        industry: str = None,
        analysis_purpose: str = 'valuation',
    ) -> Dict[str, Any]:
        """
        可比公司分析主入口 - Claude框架
        
        Args:
            target_company: 目标公司数据
            peer_companies: 可比公司列表
            industry: 行业类型
            analysis_purpose: 分析目的（valuation/efficiency/growth）
            
        Returns:
            分析结果
        """
        logger.info(f"开始可比公司分析（Claude框架）")
        
        result = {
            'target_company': target_company.get('name', 'Unknown'),
            'peer_count': len(peer_companies),
            'industry': industry,
            'analysis_purpose': analysis_purpose,
            'comps_analysis': {},
        }
        
        try:
            # Step 1: 确定分析指标（Claude: Choose the Right Metrics）
            logger.info("确定分析指标...")
            metrics = self._determine_metrics(industry, analysis_purpose)
            result['comps_analysis']['metrics'] = metrics
            
            # Step 2: 构建运营指标表（Claude: Operating Statistics）
            logger.info("构建运营指标表...")
            operating_table = self._build_operating_table(
                target_company, peer_companies, metrics['operating_metrics']
            )
            result['comps_analysis']['operating_table'] = operating_table
            
            # Step 3: 构建估值倍数表（Claude: Valuation Multiples）
            logger.info("构建估值倍数表...")
            valuation_table = self._build_valuation_table(
                target_company, peer_companies, metrics['valuation_metrics']
            )
            result['comps_analysis']['valuation_table'] = valuation_table
            
            # Step 4: 计算统计数据（Claude: Statistics Block）
            logger.info("计算统计数据...")
            operating_stats = self._calculate_statistics(operating_table, metrics['operating_metrics'])
            valuation_stats = self._calculate_statistics(valuation_table, metrics['valuation_metrics'])
            
            result['comps_analysis']['operating_statistics'] = operating_stats
            result['comps_analysis']['valuation_statistics'] = valuation_stats
            
            # Step 5: 目标公司定位分析
            logger.info("分析目标公司定位...")
            positioning = self._analyze_positioning(
                target_company, operating_table, valuation_table, operating_stats, valuation_stats
            )
            result['comps_analysis']['positioning'] = positioning
            
            # Step 6: 生成分析总结
            logger.info("生成分析总结...")
            summary = self._generate_summary(result['comps_analysis'])
            result['comps_analysis']['summary'] = summary
            
            logger.info("可比公司分析完成")
            
        except Exception as e:
            logger.error(f"可比公司分析失败: {e}")
            result['error'] = str(e)
        
        return result
    
    def _determine_metrics(
        self,
        industry: str,
        analysis_purpose: str,
    ) -> Dict[str, List[str]]:
        """
        确定分析指标（Claude: Choosing the Right Metrics）
        
        Args:
            industry: 行业类型
            analysis_purpose: 分析目的
            
        Returns:
            指标列表
        """
        # 根据行业选择指标
        industry_key = industry if industry in self.industry_metrics else 'Default'
        base_metrics = self.industry_metrics[industry_key].copy()
        
        # 根据分析目的调整指标
        if analysis_purpose == 'valuation':
            # 估值分析：重点关注估值倍数
            if 'P/E' not in base_metrics['valuation_metrics']:
                base_metrics['valuation_metrics'].append('P/E')
            if 'EV/EBITDA' not in base_metrics['valuation_metrics']:
                base_metrics['valuation_metrics'].append('EV/EBITDA')
        
        elif analysis_purpose == 'efficiency':
            # 效率分析：重点关注利润率
            efficiency_metrics = ['Gross Margin', 'EBITDA Margin', 'Net Margin', 'Asset Turnover']
            for metric in efficiency_metrics:
                if metric not in base_metrics['operating_metrics']:
                    base_metrics['operating_metrics'].append(metric)
        
        elif analysis_purpose == 'growth':
            # 增长分析：重点关注增长率
            growth_metrics = ['Revenue Growth', 'EBITDA Growth', 'Net Income Growth']
            for metric in growth_metrics:
                if metric not in base_metrics['operating_metrics']:
                    base_metrics['operating_metrics'].append(metric)
        
        return base_metrics
    
    def _build_operating_table(
        self,
        target_company: Dict[str, Any],
        peer_companies: List[Dict[str, Any]],
        metrics: List[str],
    ) -> pd.DataFrame:
        """
        构建运营指标表（Claude: Operating Statistics & Financial Metrics）
        
        Args:
            target_company: 目标公司
            peer_companies: 可比公司列表
            metrics: 指标列表
            
        Returns:
            运营指标表
        """
        # 合并所有公司
        all_companies = [target_company] + peer_companies
        
        # 构建数据表
        data = []
        for company in all_companies:
            row = {
                'Company': company.get('name', 'Unknown'),
                'Ticker': company.get('ticker', 'N/A'),
            }
            
            # 添加运营指标
            financial_data = company.get('financial_data', {})
            income_statement = financial_data.get('income_statement', {})
            balance_sheet = financial_data.get('balance_sheet', {})
            
            for metric in metrics:
                if metric == 'Revenue':
                    revenues = income_statement.get('revenue', [])
                    row['Revenue'] = revenues[-1] if revenues else 0
                
                elif metric == 'Revenue Growth':
                    growth_rates = income_statement.get('revenue_growth', [])
                    row['Revenue Growth %'] = growth_rates[-1] if growth_rates else 0
                
                elif metric == 'Gross Margin':
                    gross_margins = income_statement.get('gross_margin', [])
                    row['Gross Margin %'] = gross_margins[-1] if gross_margins else 0
                
                elif metric == 'EBITDA Margin':
                    ebitda_margins = income_statement.get('ebitda_margin', [])
                    row['EBITDA Margin %'] = ebitda_margins[-1] if ebitda_margins else 0
                
                elif metric == 'Net Margin':
                    net_margins = income_statement.get('net_margin', [])
                    row['Net Margin %'] = net_margins[-1] if net_margins else 0
                
                elif metric == 'Rule of 40':
                    # Rule of 40 = Revenue Growth % + EBITDA Margin %
                    growth = row.get('Revenue Growth %', 0)
                    margin = row.get('EBITDA Margin %', 0)
                    row['Rule of 40'] = growth + margin
                
                elif metric == 'Asset Turnover':
                    # Asset Turnover = Revenue / Total Assets
                    revenues = income_statement.get('revenue', [])
                    total_assets = balance_sheet.get('total_assets', [])
                    if revenues and total_assets and total_assets[-1] > 0:
                        row['Asset Turnover'] = revenues[-1] / total_assets[-1]
                    else:
                        row['Asset Turnover'] = 0
                
                elif metric == 'ROE':
                    # ROE = Net Income / Equity
                    net_incomes = income_statement.get('net_income', [])
                    equities = balance_sheet.get('total_equity', [])
                    if net_incomes and equities and equities[-1] > 0:
                        row['ROE %'] = (net_incomes[-1] / equities[-1]) * 100
                    else:
                        row['ROE %'] = 0
                
                elif metric == 'ROA':
                    # ROA = Net Income / Total Assets
                    net_incomes = income_statement.get('net_income', [])
                    total_assets = balance_sheet.get('total_assets', [])
                    if net_incomes and total_assets and total_assets[-1] > 0:
                        row['ROA %'] = (net_incomes[-1] / total_assets[-1]) * 100
                    else:
                        row['ROA %'] = 0
            
            data.append(row)
        
        df = pd.DataFrame(data)
        return df
    
    def _build_valuation_table(
        self,
        target_company: Dict[str, Any],
        peer_companies: List[Dict[str, Any]],
        metrics: List[str],
    ) -> pd.DataFrame:
        """
        构建估值倍数表（Claude: Valuation Multiples & Investment Metrics）
        
        Args:
            target_company: 目标公司
            peer_companies: 可比公司列表
            metrics: 指标列表
            
        Returns:
            估值倍数表
        """
        # 合并所有公司
        all_companies = [target_company] + peer_companies
        
        # 构建数据表
        data = []
        for company in all_companies:
            row = {
                'Company': company.get('name', 'Unknown'),
                'Ticker': company.get('ticker', 'N/A'),
            }
            
            # 添加估值指标
            market_data = company.get('market_data', {})
            financial_data = company.get('financial_data', {})
            income_statement = financial_data.get('income_statement', {})
            
            for metric in metrics:
                if metric == 'Market Cap':
                    row['Market Cap'] = market_data.get('market_cap', 0)
                
                elif metric == 'EV':
                    # Enterprise Value = Market Cap + Net Debt
                    market_cap = market_data.get('market_cap', 0)
                    net_debt = market_data.get('net_debt', 0)
                    row['EV'] = market_cap + net_debt
                
                elif metric == 'EV/Revenue':
                    ev = row.get('EV', 0)
                    revenues = income_statement.get('revenue', [])
                    revenue = revenues[-1] if revenues else 0
                    row['EV/Revenue'] = ev / revenue if revenue > 0 else 0
                
                elif metric == 'EV/EBITDA':
                    ev = row.get('EV', 0)
                    ebitdas = income_statement.get('ebitda', [])
                    ebitda = ebitdas[-1] if ebitdas else 0
                    row['EV/EBITDA'] = ev / ebitda if ebitda > 0 else 0
                
                elif metric == 'P/E':
                    market_cap = row.get('Market Cap', 0)
                    net_incomes = income_statement.get('net_income', [])
                    net_income = net_incomes[-1] if net_incomes else 0
                    row['P/E'] = market_cap / net_income if net_income > 0 else 0
                
                elif metric == 'P/B':
                    market_cap = row.get('Market Cap', 0)
                    equities = financial_data.get('balance_sheet', {}).get('total_equity', [])
                    equity = equities[-1] if equities else 0
                    row['P/B'] = market_cap / equity if equity > 0 else 0
                
                elif metric == 'Dividend Yield':
                    dividend_yield = market_data.get('dividend_yield', 0)
                    row['Dividend Yield %'] = dividend_yield
            
            data.append(row)
        
        df = pd.DataFrame(data)
        return df
    
    def _calculate_statistics(
        self,
        table: pd.DataFrame,
        metrics: List[str],
    ) -> Dict[str, Any]:
        """
        计算统计数据（Claude: Statistics Block）
        
        Args:
            table: 数据表
            metrics: 指标列表
            
        Returns:
            统计数据
        """
        stats = {}
        
        # 只对可比指标计算统计（排除规模指标）
        comparable_metrics = [m for m in metrics if m not in ['Revenue', 'Market Cap', 'EV']]
        
        for metric in comparable_metrics:
            # 查找对应的列名
            col_name = None
            for col in table.columns:
                if metric in col or col in metric:
                    col_name = col
                    break
            
            if col_name and col_name in table.columns:
                values = table[col_name].dropna()
                
                if len(values) > 0:
                    stats[metric] = {
                        'Maximum': values.max(),
                        '75th Percentile': np.percentile(values, 75),
                        'Median': np.median(values),
                        '25th Percentile': np.percentile(values, 25),
                        'Minimum': values.min(),
                        'Mean': values.mean(),
                        'Std Dev': values.std(),
                    }
        
        return stats
    
    def _analyze_positioning(
        self,
        target_company: Dict[str, Any],
        operating_table: pd.DataFrame,
        valuation_table: pd.DataFrame,
        operating_stats: Dict[str, Any],
        valuation_stats: Dict[str, Any],
    ) -> Dict[str, Any]:
        """
        目标公司定位分析
        
        Args:
            target_company: 目标公司
            operating_table: 运营指标表
            valuation_table: 估值倍数表
            operating_stats: 运营统计
            valuation_stats: 估值统计
            
        Returns:
            定位分析结果
        """
        positioning = {
            'valuation_position': None,
            'operating_position': None,
            'premium_discount': None,
            'analysis': '',
        }
        
        target_name = target_company.get('name', 'Unknown')
        
        # 从表中提取目标公司数据
        target_operating = operating_table[operating_table['Company'] == target_name]
        target_valuation = valuation_table[valuation_table['Company'] == target_name]
        
        if target_operating.empty or target_valuation.empty:
            return positioning
        
        # 估值定位
        if 'EV/EBITDA' in valuation_stats:
            target_ev_ebitda = target_valuation['EV/EBITDA'].values[0]
            median_ev_ebitda = valuation_stats['EV/EBITDA']['Median']
            
            if target_ev_ebitda > valuation_stats['EV/EBITDA']['75th Percentile']:
                positioning['valuation_position'] = 'Premium'
            elif target_ev_ebitda > median_ev_ebitda:
                positioning['valuation_position'] = 'Above Median'
            elif target_ev_ebitda > valuation_stats['EV/EBITDA']['25th Percentile']:
                positioning['valuation_position'] = 'Below Median'
            else:
                positioning['valuation_position'] = 'Discount'
            
            # 溢价/折价幅度
            if median_ev_ebitda > 0:
                positioning['premium_discount'] = ((target_ev_ebitda / median_ev_ebitda) - 1) * 100
        
        # 运营定位
        if 'EBITDA Margin %' in operating_stats:
            target_margin = target_operating['EBITDA Margin %'].values[0]
            median_margin = operating_stats['EBITDA Margin %']['Median']
            
            if target_margin > operating_stats['EBITDA Margin %']['75th Percentile']:
                positioning['operating_position'] = 'Top Quartile'
            elif target_margin > median_margin:
                positioning['operating_position'] = 'Above Median'
            elif target_margin > operating_stats['EBITDA Margin %']['25th Percentile']:
                positioning['operating_position'] = 'Below Median'
            else:
                positioning['operating_position'] = 'Bottom Quartile'
        
        # 生成分析文本
        positioning['analysis'] = self._generate_positioning_analysis(positioning, target_name)
        
        return positioning
    
    def _generate_positioning_analysis(
        self,
        positioning: Dict[str, Any],
        target_name: str,
    ) -> str:
        """生成定位分析文本"""
        parts = []
        
        if positioning.get('valuation_position'):
            parts.append(
                f"{target_name}估值定位：{positioning['valuation_position']}"
            )
            
            if positioning.get('premium_discount'):
                premium = positioning['premium_discount']
                if premium > 0:
                    parts.append(f"相对中值溢价{premium:.1f}%")
                else:
                    parts.append(f"相对中值折价{abs(premium):.1f}%")
        
        if positioning.get('operating_position'):
            parts.append(
                f"运营效率定位：{positioning['operating_position']}"
            )
        
        return '；'.join(parts) if parts else '定位分析数据不足'
    
    def _generate_summary(self, comps_analysis: Dict[str, Any]) -> str:
        """生成分析总结"""
        parts = []
        
        # 估值总结
        valuation_stats = comps_analysis.get('valuation_statistics', {})
        if 'EV/EBITDA' in valuation_stats:
            stats = valuation_stats['EV/EBITDA']
            parts.append(
                f"**估值倍数**：EV/EBITDA中值为{stats['Median']:.1f}x，"
                f"区间为{stats['Minimum']:.1f}x-{stats['Maximum']:.1f}x。"
            )
        
        # 运营总结
        operating_stats = comps_analysis.get('operating_statistics', {})
        if 'EBITDA Margin %' in operating_stats:
            stats = operating_stats['EBITDA Margin %']
            parts.append(
                f"**盈利能力**：EBITDA利润率中值为{stats['Median']:.1f}%，"
                f"区间为{stats['Minimum']:.1f}%-{stats['Maximum']:.1f}%。"
            )
        
        # 定位总结
        positioning = comps_analysis.get('positioning', {})
        if positioning.get('analysis'):
            parts.append(f"**定位分析**：{positioning['analysis']}")
        
        return '\n\n'.join(parts)
    
    def export_to_excel(
        self,
        result: Dict[str, Any],
        output_path: str,
    ) -> str:
        """
        导出为Excel（Claude格式）
        
        Args:
            result: 分析结果
            output_path: 输出路径
            
        Returns:
            文件路径
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            
            # Sheet 1: 运营指标
            ws1 = wb.active
            ws1.title = "Operating Metrics"
            
            # 标题
            ws1['A1'] = f"{result.get('industry', 'Industry')} - COMPARABLE COMPANY ANALYSIS"
            ws1['A1'].font = Font(bold=True, size=14, color='FFFFFF')
            ws1['A1'].fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
            ws1.merge_cells('A1:H1')
            
            # 运营指标表
            operating_table = result['comps_analysis'].get('operating_table')
            if operating_table is not None:
                # 写入数据
                for r_idx, row in enumerate(operating_table.values, start=3):
                    for c_idx, value in enumerate(row, start=1):
                        cell = ws1.cell(row=r_idx, column=c_idx, value=value)
                        
                        # 格式化
                        if c_idx > 2:  # 数值列
                            if isinstance(value, (int, float)):
                                cell.number_format = '#,##0.0' if abs(value) < 100 else '#,##0'
            
            # Sheet 2: 估值倍数
            ws2 = wb.create_sheet("Valuation Multiples")
            
            # 标题
            ws2['A1'] = f"{result.get('industry', 'Industry')} - VALUATION MULTIPLES"
            ws2['A1'].font = Font(bold=True, size=14, color='FFFFFF')
            ws2['A1'].fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
            ws2.merge_cells('A1:H1')
            
            # 估值倍数表
            valuation_table = result['comps_analysis'].get('valuation_table')
            if valuation_table is not None:
                for r_idx, row in enumerate(valuation_table.values, start=3):
                    for c_idx, value in enumerate(row, start=1):
                        cell = ws2.cell(row=r_idx, column=c_idx, value=value)
                        
                        if c_idx > 2:
                            if isinstance(value, (int, float)):
                                cell.number_format = '#,##0.0' if abs(value) < 100 else '#,##0'
            
            # Sheet 3: 统计数据
            ws3 = wb.create_sheet("Statistics")
            
            # 运营统计
            ws3['A1'] = "Operating Statistics"
            ws3['A1'].font = Font(bold=True, size=12)
            
            operating_stats = result['comps_analysis'].get('operating_statistics', {})
            row_idx = 3
            for metric, stats in operating_stats.items():
                ws3.cell(row=row_idx, column=1, value=metric)
                ws3.cell(row=row_idx, column=2, value=f"Median: {stats['Median']:.2f}")
                ws3.cell(row=row_idx, column=3, value=f"Range: {stats['Minimum']:.2f}-{stats['Maximum']:.2f}")
                row_idx += 1
            
            # 估值统计
            row_idx += 2
            ws3.cell(row=row_idx, column=1, value="Valuation Statistics")
            ws3.cell(row=row_idx, column=1).font = Font(bold=True, size=12)
            row_idx += 2
            
            valuation_stats = result['comps_analysis'].get('valuation_statistics', {})
            for metric, stats in valuation_stats.items():
                ws3.cell(row=row_idx, column=1, value=metric)
                ws3.cell(row=row_idx, column=2, value=f"Median: {stats['Median']:.2f}")
                ws3.cell(row=row_idx, column=3, value=f"Range: {stats['Minimum']:.2f}-{stats['Maximum']:.2f}")
                row_idx += 1
            
            # 保存文件
            wb.save(output_path)
            logger.info(f"Excel文件已保存: {output_path}")
            
            return output_path
            
        except Exception as e:
            logger.error(f"导出Excel失败: {e}")
            raise
